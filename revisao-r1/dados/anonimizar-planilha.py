#!/usr/bin/env python3
"""
Anonimiza Jicopa 2025.xlsx → jicopa-2025-anonimizada.xlsx

Estratégia POSICIONAL (conservadora): anonimiza apenas células de colunas
conhecidas por conter nomes de alunos. Descrições, fórmulas e códigos de
turma ficam intactos.

Mapeamento por aba:
- Alunos        : col A (>=row2) = nome completo
- Página10      : col A (>=row2) = nome; col E (>=row2) = nome (lista
                  "Não Vai")
- Descontos     : col A (>=row3) = nome (completo ou primeiro nome)
- Esportes      : col A (>=row2) = competidor. Cols 10–12 = observações
  (Futebol/Futsal/Queimada/Rouba-Bandeira/Basquete/Volêi)
                  com possíveis menções (jogador expulso etc.) — regex.
- Individual Taekwondo : col A (>=row2). Cols a partir de col 8 sob
                  rótulo "ALUNO".
- Individual Ginastica : col A (>=row2). Cols sob rótulo "ALUNA".
- Calendário, Categorias, Bandeiras : NÃO tocar (apenas códigos).

Mapa nome → "Aluno NNN" baseado na aba Alunos, ordem alfabética.
Para nomes em outras abas que não batem com a Alunos: substitui pela
máscara "Aluno (n/d)" preservando códigos de turma quando presentes.
"""
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

AQUI = Path(__file__).resolve().parent
ORIGEM = AQUI / "Jicopa 2025.xlsx"
DESTINO = AQUI / "jicopa-2025-anonimizada.xlsx"
MAPA = AQUI / ".mapeamento-anonimizacao-LOCAL.json"

TURMAS_RE = r"(?:a|b1|b2|c|d|e)\s*[mv]|gr"

PALAVRAS_NAO_NOME = {
    "ate", "nao", "sim", "veio", "pesado", "leve", "medio", "categoria",
    "infantil", "juvenil", "junior", "cadete", "mirim", "adulto",
    "feminino", "masculino", "geral", "almoco", "lanche", "turma",
    "esporte", "futebol", "futsal", "queimada", "rouba", "bandeira",
    "basquete", "volei", "taekwondo", "ginastica", "quadra", "campo",
    "salao", "matutino", "vespertino", "participa", "presente", "ausente",
    "empate", "vitoria", "perda", "todos", "individual", "coletivo",
    "colocacao", "pontos", "peso", "aluno", "aluna", "torcida", "expulso",
    "expulsa", "ata", "obs", "horario", "local", "data", "desconto",
    "descontos", "pratica", "uso", "andando", "estar", "ficar", "tempo",
    "minutos", "jogos", "jogo", "time", "times", "time1", "time2",
    "resultados", "resultado", "confrontos", "confronto", "ranking",
    "ranqueamento", "evento", "agenda", "calendario", "competidores",
    "competidor", "regras", "regulamento", "boletim", "sumula", "relatorio",
}


def normalize(s) -> str:
    if not isinstance(s, str):
        return ""
    return re.sub(r"\s+", " ", s.strip()).lower()


def sem_acento(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s)
                   if not unicodedata.combining(c))


class Anonimizador:
    def __init__(self):
        self.nome_para_codigo = {}      # nome completo normalizado → "Aluno NNN"
        self.primeiros = {}             # primeiro nome (sem acento, lower) → [códigos]
        self.subs_por_aba = {}
        self.total_subs = 0

    def carregar_mapa(self, ws_alunos):
        nomes = []
        visto = set()
        for r in range(2, ws_alunos.max_row + 1):
            v = ws_alunos.cell(r, 1).value
            if not v:
                continue
            n = normalize(v)
            if n in visto:
                continue
            visto.add(n)
            nomes.append((v.strip(), n))
        nomes.sort(key=lambda t: t[1])
        for i, (orig, n_norm) in enumerate(nomes):
            self.nome_para_codigo[n_norm] = f"Aluno {i+1:03d}"
            primeiro = sem_acento(orig.split()[0]).lower()
            self.primeiros.setdefault(primeiro, []).append(
                self.nome_para_codigo[n_norm])
        # Para o mapa local guardamos a forma humana
        self.original_para_codigo = {orig: self.nome_para_codigo[n_norm]
                                     for orig, n_norm in nomes}

    def _conta(self, aba):
        self.subs_por_aba[aba] = self.subs_por_aba.get(aba, 0) + 1
        self.total_subs += 1

    def codigo_para_nome_completo(self, valor):
        """Tenta resolver `valor` (string) para "Aluno NNN". Retorna None se não bate."""
        if not isinstance(valor, str):
            return None
        n = normalize(valor)
        if n in self.nome_para_codigo:
            return self.nome_para_codigo[n]
        return None

    def mascara_nome_isolado(self, valor, aba):
        """Substitui célula que se espera conter apenas um nome. Retorna novo valor."""
        if not isinstance(valor, str) or not valor.strip():
            return valor
        n = normalize(valor)
        if n in self.nome_para_codigo:
            self._conta(aba)
            return self.nome_para_codigo[n]
        # Primeiro nome só?
        partes = valor.strip().split()
        if len(partes) == 1:
            pn = sem_acento(partes[0]).lower()
            if pn in PALAVRAS_NAO_NOME:
                return valor
            if pn in self.primeiros and len(self.primeiros[pn]) == 1:
                self._conta(aba)
                return self.primeiros[pn][0]
            if pn in self.primeiros:
                self._conta(aba)
                return "Aluno (n/d)"
            # primeiro nome desconhecido começando com maiúscula → mascara
            if re.match(r"^[A-ZÁÉÍÓÚÂÊÔÃÕÇ][a-záéíóúâêôãõç]+$", partes[0]):
                self._conta(aba)
                return "Aluno (n/d)"
            return valor
        # 2+ palavras: tenta como nome completo, senão mascara
        # (numa coluna onde sabemos esperar nome, esse é o caso de
        #  variações ortográficas ou nomes sem acento)
        # Só mascara se a primeira palavra parecer nome próprio (maiúscula).
        if re.match(r"^[A-ZÁÉÍÓÚÂÊÔÃÕÇ]", partes[0]):
            # tentativa: primeiros nomes em ordem
            for p in partes:
                pn = sem_acento(p).lower()
                if pn in self.primeiros and len(self.primeiros[pn]) == 1:
                    self._conta(aba)
                    return self.primeiros[pn][0]
            self._conta(aba)
            return "Aluno (n/d)"
        return valor

    def mascara_observacao(self, valor, aba):
        """Aplica regex em campos de OBSERVAÇÃO (texto livre).
        Detecta padrões "Nome turma" e "Nome Sobrenome" sem destruir
        descrições de infração ou rótulos.
        """
        if not isinstance(valor, str) or not valor.strip():
            return valor
        novo = valor

        # padrão "Nome turma" (case-insensitive)
        re_nome_turma = re.compile(
            r"\b([A-Za-zÁÉÍÓÚÂÊÔÃÕÇáéíóúâêôãõç]{3,})\s+(" +
            TURMAS_RE + r")\b", re.IGNORECASE)

        def repl_nt(m):
            pn = sem_acento(m.group(1)).lower()
            tur = m.group(2)
            if pn in PALAVRAS_NAO_NOME:
                return m.group(0)
            if pn in self.primeiros and len(self.primeiros[pn]) == 1:
                self._conta(aba)
                return f"{self.primeiros[pn][0]} {tur}"
            self._conta(aba)
            return f"Aluno (n/d) {tur}"

        novo = re_nome_turma.sub(repl_nt, novo)

        # Padrão "Nome - Nome" (lista separada por hífen, ex.: "rafael -
        # gustavo - luiz Cm").
        # Substitui cada token alfabético com >=3 chars começando com letra
        # se for um primeiro nome conhecido.
        def repl_token(m):
            tok = m.group(0)
            pn = sem_acento(tok).lower()
            if pn in PALAVRAS_NAO_NOME:
                return tok
            if pn in self.primeiros and len(self.primeiros[pn]) == 1:
                self._conta(aba)
                return self.primeiros[pn][0]
            if pn in self.primeiros:
                self._conta(aba)
                return "Aluno (n/d)"
            return tok

        # Sempre roda em observações: mascara qualquer token alfabético
        # 3+ caracteres que não esteja em PALAVRAS_NAO_NOME nem comece com
        # "Aluno" (já mascarado). Preserva código de turma compacto
        # (ex.: "Cm", "Dv") e palavras técnicas/descritivas.
        def repl_token_amplo(m):
            tok = m.group(0)
            pn = sem_acento(tok).lower()
            if pn in PALAVRAS_NAO_NOME:
                return tok
            # códigos compactos de turma (Am, Bv, Cm, Dv, Em, Gr, etc.)
            if re.fullmatch(r"(?i)(a|b1|b2|c|d|e)[mv]|gr", tok):
                return tok
            # preserva "Aluno" (parte de "Aluno NNN" já posto)
            if tok.lower() == "aluno":
                return tok
            # números/sequência tipo "049" não bate na regex de letras
            # Token desconhecido alfabético: pode ser nome/apelido → mascara
            self._conta(aba)
            return "Aluno (n/d)"

        # SEMPRE roda o token-replace em observações.
        # Limiar de 5+ chars para reduzir false positives com preposições
        # ("por", "para", "com", etc.).
        novo = re.sub(r"\b[A-Za-zÁÉÍÓÚÂÊÔÃÕÇáéíóúâêôãõç]{5,}\b",
                      repl_token_amplo, novo)

        # Passe extra: primeiros nomes curtos (3-4 chars) capitalizados que
        # batem com a lista de primeiros nomes conhecida — mascara.
        def repl_curto(m):
            tok = m.group(0)
            pn = sem_acento(tok).lower()
            if pn in PALAVRAS_NAO_NOME:
                return tok
            if re.fullmatch(r"(?i)(a|b1|b2|c|d|e)[mv]|gr", tok):
                return tok
            if pn in self.primeiros:
                self._conta(aba)
                return "Aluno (n/d)"
            return tok

        novo = re.sub(
            r"\b[A-ZÁÉÍÓÚÂÊÔÃÕÇ][a-záéíóúâêôãõç]{2,3}\b",
            repl_curto, novo)

        return novo

    def processar(self, wb):
        # ---- Alunos: col A nome completo ----
        ws = wb["Alunos"]
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, 1)
            v = cell.value
            if not v:
                continue
            codigo = self.codigo_para_nome_completo(v)
            if codigo:
                cell.value = codigo
                self._conta("Alunos")
            else:
                cell.value = self.mascara_nome_isolado(v, "Alunos")

        # ---- Página10: col A e col E (lista de "Não Vai") ----
        ws = wb["Página10"]
        for r in range(2, ws.max_row + 1):
            for col in (1, 5):
                cell = ws.cell(r, col)
                v = cell.value
                if isinstance(v, str):
                    cell.value = self.mascara_nome_isolado(v, "Página10")

        # ---- Descontos: col A nome ----
        ws = wb["Descontos"]
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(r, 1)
            v = cell.value
            if isinstance(v, str):
                cell.value = self.mascara_nome_isolado(v, "Descontos")

        # ---- Esportes coletivos: col 1 competidor + cols 10-12 observações ----
        for sheet in ("Futebol", "Futsal", "Queimada", "Rouba-Bandeira",
                      "Basquete", "Volêi"):
            ws = wb[sheet]
            for r in range(2, ws.max_row + 1):
                # col 1 competidor
                cell = ws.cell(r, 1)
                if isinstance(cell.value, str):
                    cell.value = self.mascara_nome_isolado(cell.value, sheet)
                # cols 10-12 obs
                for c in range(10, 13):
                    cell = ws.cell(r, c)
                    if isinstance(cell.value, str):
                        cell.value = self.mascara_observacao(cell.value, sheet)

        # ---- Individual Taekwondo / Ginástica ----
        for sheet in ("Individual Taekwondo", "Individual Ginastica"):
            ws = wb[sheet]
            # col 1: competidor
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, 1)
                if isinstance(cell.value, str):
                    cell.value = self.mascara_nome_isolado(cell.value, sheet)
            # cols com header "ALUNO" / "ALUNA" — varre cabeçalhos
            cols_nome = set()
            for r in range(1, min(ws.max_row, 5) + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and v.strip().upper() in (
                            "ALUNO", "ALUNA"):
                        cols_nome.add(c)
            for r in range(1, ws.max_row + 1):
                for c in cols_nome:
                    cell = ws.cell(r, c)
                    if isinstance(cell.value, str) and cell.value.strip():
                        upp = cell.value.strip().upper()
                        if upp in ("ALUNO", "ALUNA"):
                            continue
                        cell.value = self.mascara_nome_isolado(
                            cell.value, sheet)


def main():
    if not ORIGEM.exists():
        sys.exit(f"ERRO: planilha de origem não encontrada em {ORIGEM}")

    print(f"[1/4] Lendo {ORIGEM.name} ...")
    wb = openpyxl.load_workbook(ORIGEM)

    a = Anonimizador()
    a.carregar_mapa(wb["Alunos"])
    print(f"[1/4] {len(a.nome_para_codigo)} nomes únicos identificados.")

    print("[2/4] Processando abas posicionalmente...")
    a.processar(wb)
    print(f"[2/4] {a.total_subs} substituições aplicadas.")
    for k, v in a.subs_por_aba.items():
        print(f"        {k}: {v}")

    print(f"[3/4] Salvando {DESTINO.name} ...")
    wb.save(DESTINO)

    print(f"[4/4] Salvando mapa LOCAL em {MAPA.name} ...")
    mapa_completo = {
        "_aviso": ("ARQUIVO PRIVADO. NÃO versionar. Permite "
                   "re-identificação dos alunos da planilha 2025."),
        "ordem": "alfabetica_case_insensitive",
        "total_alunos": len(a.nome_para_codigo),
        "mapa": a.original_para_codigo,
    }
    MAPA.write_text(json.dumps(mapa_completo, ensure_ascii=False, indent=2),
                    encoding="utf-8")

    print("\nOK.")
    print(f"   Anonimizada: {DESTINO}")
    print(f"   Mapa local:  {MAPA} (NÃO versionado)")


if __name__ == "__main__":
    main()
