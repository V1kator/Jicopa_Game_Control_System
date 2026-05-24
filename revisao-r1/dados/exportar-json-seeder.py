#!/usr/bin/env python3
"""
Lê dados/jicopa-2025-anonimizada.xlsx e gera dados/jicopa-2025-dados.json
no formato consumido por jicopa/database/seeders/Jicopa2025Seeder.php.

Estrutura do JSON:
{
  "meta": {...},
  "turmas":     [{"name", "period"}, ...],
  "categorias": [{"name", "turmas": [...], "esportes": [...]}, ...],
  "esportes":   [{"name", "type"}, ...],
  "alunos":     [{"name", "turma", "period", "esportes": [...]}, ...],
  "jogos":      [{"categoria", "esporte", "time1", "time2", "data",
                 "hora", "local", "vencedor"}, ...],
  "penalidades":[{"turma_codigo", "motivo", "pontos"}, ...],
  "bandeiras":  {
    "config": {"num_jurados", "nota_min", "nota_max", "bonus"},
    "notas":  [{"turma_codigo", "categoria", "notas": [n1,n2,n3,n4]}, ...]
  }
}
"""
import json
import re
import sys
from pathlib import Path
from datetime import date

import openpyxl

AQUI = Path(__file__).resolve().parent
PLANILHA = AQUI / "jicopa-2025-anonimizada.xlsx"
SAIDA = AQUI / "jicopa-2025-dados.json"

# Mapeamento turma_codigo (como aparece na planilha) → (nome, período)
TURMA_CODIGO_PARA_NOME = {
    "A m":  ("Turma A",  "Matutino"),
    "A v":  ("Turma A",  "Vespertino"),
    "B1 m": ("Turma B1", "Matutino"),
    "B1 v": ("Turma B1", "Vespertino"),
    "B2 m": ("Turma B2", "Matutino"),
    "B2 v": ("Turma B2", "Vespertino"),
    "C m":  ("Turma C",  "Matutino"),
    "C v":  ("Turma C",  "Vespertino"),
    "D m":  ("Turma D",  "Matutino"),
    "D v":  ("Turma D",  "Vespertino"),
    "E m":  ("Turma E",  "Matutino"),
    "E v":  ("Turma E",  "Vespertino"),
    "GR":   ("Turma GR", "Vespertino"),
}

# Mapeamento categoria → turmas (conforme aba Categorias da planilha)
CATEGORIA_TURMAS = {
    "Infantil": ["A m", "A v", "B1 m", "B1 v", "B2 m", "B2 v"],
    "Juvenil":  ["C m", "C v", "D m", "D v"],
    "Júnior":   ["E m", "E v", "GR"],
}

# Esportes e tipo
ESPORTES = [
    ("Futebol",        "coletivo"),
    ("Futsal",         "coletivo"),
    ("Queimada",       "coletivo"),
    ("Rouba Bandeira", "coletivo"),
    ("Basquete",       "coletivo"),
    ("Vôlei",          "coletivo"),
    ("Taekwondo",      "individual"),
    ("Ginástica",      "individual"),
]

# Mapeamento de nome de aba de esporte → nome canônico do esporte
ABA_PARA_ESPORTE = {
    "Futebol":        "Futebol",
    "Futsal":         "Futsal",
    "Queimada":       "Queimada",
    "Rouba-Bandeira": "Rouba Bandeira",
    "Basquete":       "Basquete",
    "Volêi":          "Vôlei",
}

# Datas plausíveis para 2 jogos por esporte (semana 19-23 de maio de 2025).
# Não tentamos reproduzir o calendário 2025 célula a célula porque a estrutura
# da aba Calendário é heterogênea; usamos um mini-calendário compacto.
DATAS_JOGOS = [
    (date(2025, 5, 19), "08:00:00", "Quadra"),  # Segunda
    (date(2025, 5, 20), "09:30:00", "Quadra"),  # Terça
    (date(2025, 5, 21), "08:00:00", "Campo"),   # Quarta
    (date(2025, 5, 22), "09:30:00", "Campo"),   # Quinta
    (date(2025, 5, 23), "08:00:00", "Quadra"),  # Sexta
]


def normalize_turma(s):
    """Converte 'a v', 'A V', 'B2 M', 'GR' em código canônico ('A v', 'B2 m', 'GR')."""
    if not isinstance(s, str):
        return None
    s = re.sub(r"\s+", " ", s.strip())
    if not s:
        return None
    up = s.upper()
    if up == "GR":
        return "GR"
    m = re.match(r"^(A|B1|B2|C|D|E)\s*([MV])$", up)
    if m:
        letra = m.group(1)
        per = "m" if m.group(2) == "M" else "v"
        return f"{letra} {per}"
    return None


def categoria_da_turma(turma_cod):
    for cat, lista in CATEGORIA_TURMAS.items():
        if turma_cod in lista:
            return cat
    return None


def extrair_alunos(ws):
    """Lê a aba Alunos e devolve lista de dicts."""
    header = [str(ws.cell(1, c).value or "").strip() for c in range(1, 12)]
    # Espera header[0]=Nome, 1=Turma, 2=Período, 3..10=esportes
    nome_col = 1
    turma_col = 2
    periodo_col = 3
    esporte_cols = list(range(4, 12))  # 4..11 = 8 esportes
    esporte_nomes = [header[c - 1] for c in esporte_cols]
    # normaliza nomes de coluna para baterem com ESPORTES (lista canônica)
    norm = {
        "Futebol": "Futebol",
        "Futsal": "Futsal",
        "Queimada": "Queimada",
        "Rouba Bandeira": "Rouba Bandeira",
        "Rouba-Bandeira": "Rouba Bandeira",
        "Basquete": "Basquete",
        "Vôlei": "Vôlei",
        "Volêi": "Vôlei",
        "Taekwondo": "Taekwondo",
        "Ginastica": "Ginástica",
        "Ginástica": "Ginástica",
    }
    esporte_canon = [norm.get(n.strip(), n.strip()) for n in esporte_nomes]

    alunos = []
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(r, nome_col).value
        if not nome:
            continue
        turma_raw = ws.cell(r, turma_col).value
        per = ws.cell(r, periodo_col).value
        if not turma_raw or not per:
            continue
        turma_nome = str(turma_raw).strip()  # ex: "Turma E"
        periodo = str(per).strip()
        esportes_aluno = []
        for idx, c in enumerate(esporte_cols):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == "participa":
                esportes_aluno.append(esporte_canon[idx])
        alunos.append({
            "name": str(nome).strip(),
            "turma": turma_nome,
            "period": periodo,
            "esportes": esportes_aluno,
        })
    return alunos


def extrair_jogos(wb):
    """Para cada aba de esporte coletivo, pega os 2 primeiros confrontos."""
    jogos = []
    idx_data = 0
    for aba, esporte in ABA_PARA_ESPORTE.items():
        ws = wb[aba]
        # Os confrontos estão em col 6 (Categoria), 7 (Time1), 8 (Time2),
        # 9 (Resultado). A primeira linha é cabeçalho "CONFRONTOS", a 2ª é
        # "Categoria | Turmas | | Resultados", e os dados começam em r >= 3.
        confrontos_aba = []
        categoria_atual = None
        for r in range(3, min(ws.max_row, 60) + 1):
            cat_v = ws.cell(r, 6).value
            if cat_v and str(cat_v).strip():
                categoria_atual = str(cat_v).strip()
            t1 = normalize_turma(ws.cell(r, 7).value)
            t2 = normalize_turma(ws.cell(r, 8).value)
            res = ws.cell(r, 9).value
            if t1 and t2:
                vencedor = None
                if isinstance(res, str):
                    rt = res.strip().lower()
                    if rt == "empate":
                        vencedor = "empate"
                    else:
                        # Resultado é uma das turmas
                        cand = normalize_turma(res)
                        if cand:
                            vencedor = cand
                confrontos_aba.append({
                    "categoria": categoria_atual,
                    "time1": t1,
                    "time2": t2,
                    "vencedor": vencedor,
                })
                if len(confrontos_aba) >= 2:
                    break
        for conf in confrontos_aba[:2]:
            data, hora, local = DATAS_JOGOS[idx_data % len(DATAS_JOGOS)]
            idx_data += 1
            jogos.append({
                "esporte": esporte,
                "categoria": conf["categoria"] or categoria_da_turma(conf["time1"]) or "Infantil",
                "time1": conf["time1"],
                "time2": conf["time2"],
                "vencedor": conf["vencedor"],
                "data": data.isoformat(),
                "hora": hora,
                "local": local,
                "placar_time1": None,  # planilha não tem placar numérico
                "placar_time2": None,
            })

    # Esportes individuais: criar 2 "jogos" representando 2 sessões da
    # competição. time1/time2 ficam null (NULL no banco).
    for esp_aba, esp_nome in [("Individual Taekwondo", "Taekwondo"),
                              ("Individual Ginastica", "Ginástica")]:
        for i in range(2):
            data, hora, local = DATAS_JOGOS[idx_data % len(DATAS_JOGOS)]
            idx_data += 1
            jogos.append({
                "esporte": esp_nome,
                "categoria": "Infantil",  # representativa
                "time1": None,
                "time2": None,
                "vencedor": None,
                "data": data.isoformat(),
                "hora": hora,
                "local": local,
                "placar_time1": None,
                "placar_time2": None,
            })
    return jogos


def extrair_penalidades(ws):
    """Aba Descontos: cols A=nome (anonimizado), B=turma, C=motivo, D=pontos."""
    pen = []
    for r in range(3, ws.max_row + 1):
        nome = ws.cell(r, 1).value
        turma_raw = ws.cell(r, 2).value
        motivo = ws.cell(r, 3).value
        qtd = ws.cell(r, 4).value
        if not motivo:
            continue
        try:
            pontos = int(qtd)
        except (TypeError, ValueError):
            continue
        turma_cod = normalize_turma(turma_raw)
        is_turma_inteira = (isinstance(turma_raw, str)
                            and turma_raw.strip().lower() == "todos")
        pen.append({
            "tipo_destino": "turma" if (turma_cod or is_turma_inteira) else "aluno",
            "turma_codigo": turma_cod,
            "todos": is_turma_inteira,
            "motivo": str(motivo).strip(),
            "pontos": pontos,
        })
    return pen


def extrair_bandeiras(ws):
    """Aba Bandeiras: A=Turma, B-E=4 jurados, F=Total."""
    notas = []
    for r in range(2, ws.max_row + 1):
        turma_raw = ws.cell(r, 1).value
        if not turma_raw:
            continue
        turma_cod = normalize_turma(turma_raw)
        if not turma_cod:
            continue
        n1 = ws.cell(r, 2).value
        n2 = ws.cell(r, 3).value
        n3 = ws.cell(r, 4).value
        n4 = ws.cell(r, 5).value
        if not all(isinstance(x, (int, float)) for x in (n1, n2, n3, n4)):
            continue
        notas.append({
            "turma_codigo": turma_cod,
            "categoria": categoria_da_turma(turma_cod) or "Infantil",
            "notas": [float(n1), float(n2), float(n3), float(n4)],
        })
    return notas


def main():
    if not PLANILHA.exists():
        sys.exit(f"ERRO: planilha anonimizada não encontrada em {PLANILHA}")

    print(f"[1/6] Lendo {PLANILHA.name} ...")
    wb = openpyxl.load_workbook(PLANILHA, data_only=True)

    print("[2/6] Extraindo alunos...")
    alunos = extrair_alunos(wb["Alunos"])
    # Dedup: a planilha tem 1 duplicata (caso variando para 274 únicos).
    vistos = set()
    alunos_unicos = []
    for a in alunos:
        chave = (a["name"], a["turma"], a["period"])
        if chave in vistos:
            continue
        vistos.add(chave)
        alunos_unicos.append(a)
    if len(alunos) != len(alunos_unicos):
        print(f"        {len(alunos)} linhas → {len(alunos_unicos)} alunos "
              f"únicos (descartadas {len(alunos) - len(alunos_unicos)} "
              f"duplicatas).")
    else:
        print(f"        {len(alunos_unicos)} alunos.")
    alunos = alunos_unicos

    print("[3/6] Inferindo turmas únicas...")
    turmas_set = set()
    for a in alunos:
        turmas_set.add((a["turma"], a["period"]))
    turmas = [{"name": n, "period": p} for n, p in sorted(turmas_set)]
    print(f"        {len(turmas)} turmas únicas.")

    print("[4/6] Extraindo jogos representativos (2/esporte)...")
    jogos = extrair_jogos(wb)
    print(f"        {len(jogos)} jogos.")

    print("[5/6] Extraindo penalidades e bandeiras...")
    penalidades = extrair_penalidades(wb["Descontos"])
    bandeiras_notas = extrair_bandeiras(wb["Bandeiras"])
    print(f"        {len(penalidades)} penalidades; "
          f"{len(bandeiras_notas)} bandeiras.")

    # Categorias
    categorias = []
    for cat, turmas_codes in CATEGORIA_TURMAS.items():
        categorias.append({
            "name": cat,
            "turmas": turmas_codes,
            "esportes": [e[0] for e in ESPORTES],  # cada cat compete em todos
        })

    saida = {
        "meta": {
            "origem": "jicopa-2025-anonimizada.xlsx",
            "gerado_em": "2026-05-22",
            "total_alunos": len(alunos),
            "total_jogos": len(jogos),
            "total_penalidades": len(penalidades),
            "total_bandeiras": len(bandeiras_notas),
        },
        "turmas": turmas,
        "categorias": categorias,
        "esportes": [{"name": n, "type": t} for n, t in ESPORTES],
        "alunos": alunos,
        "jogos": jogos,
        "penalidades": penalidades,
        "bandeiras": {
            "config": {
                "num_jurados": 4,
                "nota_min": 0.0,
                "nota_max": 300.0,
                "bonus": 5,
            },
            "notas": bandeiras_notas,
        },
    }

    print(f"[6/6] Salvando {SAIDA.name} ...")
    SAIDA.write_text(json.dumps(saida, ensure_ascii=False, indent=2),
                     encoding="utf-8")
    tam_kb = SAIDA.stat().st_size / 1024
    print(f"\nOK. {SAIDA} ({tam_kb:.1f} KB)")


if __name__ == "__main__":
    main()
